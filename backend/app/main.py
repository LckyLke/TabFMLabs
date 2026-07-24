"""FastAPI app: upload a table, mark it on the grid, predict empty target cells."""

import io
import json
import logging
import os
import threading
import uuid
from contextlib import asynccontextmanager

import numpy as np
import pandas as pd
from fastapi import FastAPI, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import Response, StreamingResponse
from sklearn.metrics import (
    accuracy_score,
    confusion_matrix,
    f1_score,
    mean_absolute_error,
    r2_score,
)
from sklearn.model_selection import train_test_split

from . import datasets, store
from .inference import (
    DEFAULT_MODEL,
    MAX_CLASSES,
    MODELS,
    FitPredictResult,
    WeightsUnavailable,
    get_backend,
    infer_task,
    resolve_model,
)
from .schemas import (
    ConfusionMatrix,
    CvCheckResponse,
    DatasetResponse,
    DistributionBin,
    ExplainRequest,
    ExplainResponse,
    ExplainRowRequest,
    ExplainRowResponse,
    FeatureImportance,
    GridPage,
    HoldoutSample,
    ImputedColumn,
    ImputeRequest,
    ImputeResponse,
    JobStatus,
    Metrics,
    ModelInfo,
    PredictionRow,
    PredictRequest,
    PredictResponse,
    ProjectInfo,
    RolesRequest,
    RowContribution,
    SheetRequest,
    TableProfile,
    TableSpec,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

MIN_CONTEXT_ROWS = 5
MIN_ROWS_FOR_HOLDOUT = 20
MIN_RELIABLE_HOLDOUT = 10
HOLDOUT_FRACTION = 0.2
CV_FOLDS = 5
MAX_RESULT_ROWS_INLINE = 500
MAX_HOLDOUT_SAMPLES = 200

@asynccontextmanager
async def _lifespan(_: FastAPI):
    removed = store.dedupe_projects()
    if removed:
        logger.info("Removed %d duplicate project(s)", removed)
    yield


app = FastAPI(title="TabFM Studio API", lifespan=_lifespan)

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://127.0.0.1:5173",
        *[o for o in os.environ.get("EXTRA_CORS_ORIGINS", "").split(",") if o],
    ],
    allow_methods=["*"],
    allow_headers=["*"],
)


# --------------------------------------------------------------------------
# Datasets & projects
# --------------------------------------------------------------------------


def _profile(df) -> TableProfile:
    columns = datasets.profile_columns(df)
    return TableProfile(
        n_rows=len(df),
        columns=columns,
        warnings=datasets.dataset_warnings(df, columns),
    )


def _dataset_response(dataset_id: str, ds: datasets.Dataset) -> DatasetResponse:
    sheet = ds.sheet
    grid, truncated = datasets.grid_rows(sheet.raw)
    return DatasetResponse(
        dataset_id=dataset_id,
        filename=ds.filename,
        sheet_names=list(ds.sheets.keys()),
        active_sheet=ds.active,
        n_raw_rows=len(sheet.raw),
        grid=grid,
        grid_truncated=truncated,
        spec=sheet.spec,
        table=_profile(sheet.df),
        roles=sheet.roles,
    )


@app.post("/api/datasets", response_model=DatasetResponse)
async def upload_dataset(file: UploadFile) -> DatasetResponse:
    content = await file.read()
    max_bytes = int(os.environ.get("MAX_UPLOAD_BYTES", "0"))
    if max_bytes and len(content) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"This demo instance accepts files up to {max_bytes:,} bytes.",
        )
    filename = file.filename or "upload.csv"
    try:
        raws = datasets.parse_upload(filename, content)
        dataset_id, ds = datasets.store_dataset(filename, raws, content)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return _dataset_response(dataset_id, ds)


@app.get("/api/datasets/{dataset_id}", response_model=DatasetResponse)
def get_dataset(dataset_id: str) -> DatasetResponse:
    try:
        ds = datasets.get_dataset(dataset_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Dataset not found.")
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return _dataset_response(dataset_id, ds)


@app.get("/api/datasets/{dataset_id}/grid", response_model=GridPage)
def get_grid_page(dataset_id: str, offset: int = 0, limit: int = 500) -> GridPage:
    try:
        ds = datasets.get_dataset(dataset_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Dataset not found.")
    limit = max(1, min(limit, 2000))
    offset = max(0, offset)
    return GridPage(
        offset=offset,
        rows=datasets.grid_page(ds.sheet.raw, offset, limit),
        n_raw_rows=len(ds.sheet.raw),
    )


@app.put("/api/datasets/{dataset_id}/sheet", response_model=DatasetResponse)
def set_active_sheet(dataset_id: str, req: SheetRequest) -> DatasetResponse:
    try:
        ds = datasets.set_active_sheet(dataset_id, req.sheet_name)
    except KeyError:
        raise HTTPException(status_code=404, detail="Dataset not found. Re-upload the file.")
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return _dataset_response(dataset_id, ds)


@app.put("/api/datasets/{dataset_id}/spec", response_model=TableProfile)
def set_table_spec(dataset_id: str, spec: TableSpec) -> TableProfile:
    try:
        ds = datasets.update_spec(dataset_id, spec)
    except KeyError:
        raise HTTPException(status_code=404, detail="Dataset not found. Re-upload the file.")
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return _profile(ds.sheet.df)


@app.put("/api/datasets/{dataset_id}/roles")
def set_roles(dataset_id: str, req: RolesRequest) -> dict:
    try:
        datasets.set_roles(dataset_id, req.roles)
    except KeyError:
        raise HTTPException(status_code=404, detail="Dataset not found.")
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    return {"ok": True}


@app.get("/api/projects", response_model=list[ProjectInfo])
def list_projects() -> list[ProjectInfo]:
    return [
        ProjectInfo(
            dataset_id=p["dataset_id"],
            filename=p["filename"],
            created_at=p["created_at"],
            last_used=p["last_used"],
            n_results=p["n_results"],
            target_column=p["target_column"],
            has_result=p["n_results"] > 0,
        )
        for p in store.list_projects()
    ]


@app.delete("/api/projects/{dataset_id}")
def delete_project(dataset_id: str) -> dict:
    if not store.delete_project(dataset_id):
        raise HTTPException(status_code=404, detail="Project not found.")
    datasets._datasets.pop(dataset_id, None)
    return {"ok": True}


# --------------------------------------------------------------------------
# Models
# --------------------------------------------------------------------------


@app.get("/api/models", response_model=list[ModelInfo])
def list_models() -> list[ModelInfo]:
    return [
        ModelInfo(
            id=model_id,
            label=info["label"],
            description=info["description"],
            is_default=model_id == DEFAULT_MODEL,
            supports_ensemble=info["supports_ensemble"],
        )
        for model_id, info in MODELS.items()
    ]


# --------------------------------------------------------------------------
# Prediction core (shared by the sync endpoint and background jobs)
# --------------------------------------------------------------------------


def _prepare(req: PredictRequest):
    try:
        ds = datasets.get_dataset(req.dataset_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Dataset not found. Re-upload the file.")
    df = ds.sheet.df

    missing = [c for c in [req.target_column, *req.feature_columns] if c not in df.columns]
    if missing:
        raise HTTPException(status_code=422, detail=f"Unknown columns: {missing}")
    if not req.feature_columns:
        raise HTTPException(status_code=422, detail="Select at least one feature column.")
    if req.target_column in req.feature_columns:
        raise HTTPException(status_code=422, detail="The target column cannot also be a feature.")

    target = df[req.target_column]
    features = df[req.feature_columns]
    labeled_mask = target.notna()
    X_context, y_context = features[labeled_mask], target[labeled_mask]
    X_predict = features[~labeled_mask]

    if len(X_predict) == 0:
        raise HTTPException(
            status_code=422,
            detail=f"No rows to predict: every row has a value in '{req.target_column}'. "
            "Leave the target cell empty for rows you want predicted.",
        )
    if len(X_context) < MIN_CONTEXT_ROWS:
        raise HTTPException(
            status_code=422,
            detail=f"Only {len(X_context)} labeled rows found; at least "
            f"{MIN_CONTEXT_ROWS} are needed as examples for the model.",
        )

    task = req.task or infer_task(target)
    if task == "classification" and y_context.nunique() > MAX_CLASSES:
        raise HTTPException(
            status_code=422,
            detail=f"'{req.target_column}' has {y_context.nunique()} distinct values; "
            f"at most {MAX_CLASSES} classes are supported. "
            + (
                "Switch the task to “Numbers (regression)” instead."
                if req.task == "classification"
                else "For a numeric target this would be treated as regression automatically."
            ),
        )
    if task == "regression" and not pd.api.types.is_numeric_dtype(y_context):
        raise HTTPException(
            status_code=422,
            detail=f"'{req.target_column}' contains non-numeric values, so it cannot be "
            "predicted as a number. Use “Categories (classification)” instead.",
        )

    try:
        model_id = resolve_model(req.model)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if req.ensemble and not MODELS[model_id]["supports_ensemble"]:
        raise HTTPException(
            status_code=422,
            detail=f"{MODELS[model_id]['label']} does not support ensemble mode.",
        )
    backend = get_backend(model_id)

    return ds, df, labeled_mask, X_context, y_context, X_predict, task, backend


def _subsample(X, y, max_rows: int | None, warnings: list[str]):
    if max_rows is None or len(X) <= max_rows:
        return X, y
    idx = X.sample(n=max_rows, random_state=42).index.sort_values()
    warnings.append(
        f"Using a random sample of {max_rows:,} of the {len(X):,} example rows as context."
    )
    return X.loc[idx], y.loc[idx]


def _run_predict(req: PredictRequest, progress=lambda stage: None) -> PredictResponse:
    ds, df, labeled_mask, X_context, y_context, X_predict, task, backend = _prepare(req)
    warnings: list[str] = []
    X_context, y_context = _subsample(X_context, y_context, req.max_context_rows, warnings)

    try:
        progress("accuracy check")
        metrics = _holdout_metrics(
            backend, X_context, y_context, task, warnings, ensemble=req.ensemble
        )
        progress(f"predicting {len(X_predict)} rows")
        result = backend.fit_predict(X_context, y_context, X_predict, task, ensemble=req.ensemble)
    except WeightsUnavailable as exc:
        raise HTTPException(
            status_code=503,
            detail=f"The {MODELS.get(req.model or DEFAULT_MODEL, {}).get('label', 'selected')} "
            f"model is not available yet: {exc}",
        ) from exc
    except HTTPException:
        raise
    except Exception as exc:  # surface model errors as a readable message
        logger.exception("Prediction failed")
        raise HTTPException(status_code=500, detail=f"Model inference failed: {exc}") from exc

    progress("preparing results")
    predictions = result.predictions
    if task == "regression":
        # 6 significant digits: enough precision, no float-noise tails
        predictions = np.asarray([float(f"{float(v):.6g}") for v in predictions])
        result = FitPredictResult(predictions, result.confidences, result.model_name)

    result_df = df.copy()
    pred_col = f"{req.target_column} (predicted)"
    result_df[pred_col] = pd.NA
    result_df.loc[~labeled_mask, pred_col] = result.predictions
    if result.confidences is not None:
        result_df["confidence"] = pd.NA
        result_df.loc[~labeled_mask, "confidence"] = np.round(result.confidences, 3)

    rows = _prediction_rows(df, labeled_mask, result)
    all_rows = rows
    if len(rows) > MAX_RESULT_ROWS_INLINE:
        warnings.append(
            f"Showing the first {MAX_RESULT_ROWS_INLINE} of {len(rows)} predicted rows in "
            "the grid preview. The download contains all of them."
        )
        rows = rows[:MAX_RESULT_ROWS_INLINE]

    prediction_id = uuid.uuid4().hex[:12]
    response = PredictResponse(
        prediction_id=prediction_id,
        task=task,
        model_name=result.model_name,
        n_context=len(X_context),
        n_predicted=len(X_predict),
        metrics=metrics,
        predictions=rows,
        distribution=_distribution(result.predictions, task),
        warnings=warnings,
    )

    csv_buf = io.StringIO()
    result_df.to_csv(csv_buf, index=False)
    persisted = response.model_copy(update={"predictions": all_rows})
    store.save_result(
        prediction_id,
        req.dataset_id,
        ds.active,
        req.target_column,
        persisted.model_dump_json(),
        csv_buf.getvalue().encode(),
    )
    return response


@app.post("/api/predict", response_model=PredictResponse)
def predict(req: PredictRequest) -> PredictResponse:
    return _run_predict(req)


# --------------------------------------------------------------------------
# Prediction jobs (progress + cancel)
# --------------------------------------------------------------------------

_jobs: dict[str, dict] = {}
_jobs_lock = threading.Lock()


@app.post("/api/predict-jobs", response_model=JobStatus)
def start_predict_job(req: PredictRequest) -> JobStatus:
    # Validate cheap things up-front so obvious errors are synchronous.
    _prepare(req)
    job_id = uuid.uuid4().hex[:12]
    job = {"status": "queued", "stage": "queued", "result": None, "error": None, "error_code": None, "cancelled": False}
    with _jobs_lock:
        _jobs[job_id] = job

    def progress(stage: str) -> None:
        if job["cancelled"]:
            raise _JobCancelled()
        job["stage"] = stage

    def run() -> None:
        job["status"] = "running"
        try:
            job["result"] = _run_predict(req, progress)
            job["status"] = "cancelled" if job["cancelled"] else "done"
        except _JobCancelled:
            job["status"] = "cancelled"
        except HTTPException as exc:
            job["status"] = "error"
            job["error"] = exc.detail
            job["error_code"] = exc.status_code
        except Exception as exc:  # pragma: no cover
            logger.exception("Job failed")
            job["status"] = "error"
            job["error"] = str(exc)
            job["error_code"] = 500

    threading.Thread(target=run, daemon=True).start()
    return _job_status(job_id, job)


class _JobCancelled(Exception):
    pass


def _job_status(job_id: str, job: dict) -> JobStatus:
    return JobStatus(
        job_id=job_id,
        status=job["status"],
        stage=job["stage"],
        result=job["result"] if job["status"] == "done" else None,
        error=job["error"],
        error_code=job["error_code"],
    )


@app.get("/api/predict-jobs/{job_id}", response_model=JobStatus)
def get_predict_job(job_id: str) -> JobStatus:
    with _jobs_lock:
        job = _jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found.")
    return _job_status(job_id, job)


@app.delete("/api/predict-jobs/{job_id}", response_model=JobStatus)
def cancel_predict_job(job_id: str) -> JobStatus:
    with _jobs_lock:
        job = _jobs.get(job_id)
    if job is None:
        raise HTTPException(status_code=404, detail="Job not found.")
    job["cancelled"] = True
    if job["status"] == "queued":
        job["status"] = "cancelled"
    return _job_status(job_id, job)


# --------------------------------------------------------------------------
# Impute (fill every empty cell, one pass per incomplete column)
# --------------------------------------------------------------------------


@app.post("/api/impute", response_model=ImputeResponse)
def impute(req: ImputeRequest) -> ImputeResponse:
    try:
        ds = datasets.get_dataset(req.dataset_id)
    except KeyError:
        raise HTTPException(status_code=404, detail="Dataset not found. Re-upload the file.")
    df = ds.sheet.df

    try:
        model_id = resolve_model(req.model)
    except ValueError as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc
    if req.ensemble and not MODELS[model_id]["supports_ensemble"]:
        raise HTTPException(
            status_code=422,
            detail=f"{MODELS[model_id]['label']} does not support ensemble mode.",
        )
    backend = get_backend(model_id)

    roles = ds.sheet.roles or ["feature"] * len(df.columns)
    usable = [c for c, role in zip(df.columns, roles) if role != "ignore"]
    incomplete = [c for c in usable if df[c].isna().any()]
    if not incomplete:
        raise HTTPException(
            status_code=422, detail="No empty cells to fill in the marked columns."
        )

    warnings: list[str] = []
    result_df = df.copy()
    filled: list[ImputedColumn] = []
    model_name = getattr(backend, "name", "model")

    for col in incomplete:
        labeled = df[col].notna()
        if int(labeled.sum()) < MIN_CONTEXT_ROWS:
            warnings.append(f"“{col}”: fewer than {MIN_CONTEXT_ROWS} filled rows — skipped.")
            continue
        y = df[col][labeled]
        task = infer_task(y)
        if task == "classification" and y.nunique() > MAX_CLASSES:
            warnings.append(f"“{col}”: more than {MAX_CLASSES} classes — skipped.")
            continue
        features = [c for c in usable if c != col]
        X_context, y_context = _subsample(
            df.loc[labeled, features], y, req.max_context_rows, warnings
        )
        try:
            result = backend.fit_predict(
                X_context, y_context, df.loc[~labeled, features], task, ensemble=req.ensemble
            )
        except WeightsUnavailable as exc:
            raise HTTPException(status_code=503, detail=str(exc)) from exc
        except Exception:
            logger.exception("Imputing %s failed", col)
            warnings.append(f"“{col}”: the model failed — skipped.")
            continue
        predictions = result.predictions
        if task == "regression":
            predictions = np.asarray([float(f"{float(v):.6g}") for v in predictions])
        result_df.loc[~labeled, col] = predictions
        filled.append(ImputedColumn(column=col, task=task, n_filled=int((~labeled).sum())))
        model_name = result.model_name

    if not filled:
        raise HTTPException(
            status_code=422,
            detail="No column could be filled. " + " ".join(warnings),
        )

    prediction_id = uuid.uuid4().hex[:12]
    response = ImputeResponse(
        prediction_id=prediction_id,
        model_name=model_name,
        columns=filled,
        n_cells_filled=sum(c.n_filled for c in filled),
        warnings=warnings,
    )
    csv_buf = io.StringIO()
    result_df.to_csv(csv_buf, index=False)
    # "predictions": [] keeps the xlsx download route harmless for impute results.
    stored = {**response.model_dump(), "predictions": []}
    store.save_result(
        prediction_id,
        req.dataset_id,
        ds.active,
        filled[0].column,
        json.dumps(stored),
        csv_buf.getvalue().encode(),
    )
    return response


# --------------------------------------------------------------------------
# Thorough check (k-fold cross-validation)
# --------------------------------------------------------------------------


@app.post("/api/cv-check", response_model=CvCheckResponse)
def cv_check(req: PredictRequest) -> CvCheckResponse:
    _, _, _, X, y, _, task, backend = _prepare(req)
    warnings: list[str] = []
    X, y = _subsample(X, y, req.max_context_rows, warnings)
    if len(X) < MIN_ROWS_FOR_HOLDOUT:
        raise HTTPException(
            status_code=422,
            detail=f"The thorough check needs at least {MIN_ROWS_FOR_HOLDOUT} labeled rows.",
        )

    from sklearn.model_selection import KFold, StratifiedKFold

    stratified = task == "classification" and y.value_counts().min() >= CV_FOLDS
    if stratified:
        splitter = StratifiedKFold(n_splits=CV_FOLDS, shuffle=True, random_state=42)
    else:
        splitter = KFold(n_splits=CV_FOLDS, shuffle=True, random_state=42)

    scores: list[float] = []
    try:
        for train_idx, val_idx in splitter.split(X, y if stratified else None):
            X_tr, y_tr = X.iloc[train_idx], y.iloc[train_idx]
            X_va, y_va = X.iloc[val_idx], y.iloc[val_idx]
            result = backend.fit_predict(X_tr, y_tr, X_va, task, ensemble=req.ensemble)
            if task == "classification":
                scores.append(float(accuracy_score(y_va, result.predictions)))
            else:
                scores.append(float(r2_score(y_va, result.predictions)))
    except WeightsUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc

    return CvCheckResponse(
        task=task,
        model_name=getattr(backend, "name", "model"),
        metric_name="accuracy" if task == "classification" else "R²",
        n_folds=CV_FOLDS,
        n_labeled=len(X),
        scores=[round(s, 4) for s in scores],
        mean=round(float(np.mean(scores)), 4),
        std=round(float(np.std(scores)), 4),
    )


# --------------------------------------------------------------------------
# Explain (permutation importance)
# --------------------------------------------------------------------------


@app.post("/api/explain", response_model=ExplainResponse)
def explain(req: ExplainRequest) -> ExplainResponse:
    predict_req = PredictRequest(
        dataset_id=req.dataset_id,
        target_column=req.target_column,
        feature_columns=req.feature_columns,
        model=req.model,
        task=req.task,
        ensemble=req.ensemble,
    )
    _, _, _, X, y, _, task, backend = _prepare(predict_req)
    if len(X) < MIN_ROWS_FOR_HOLDOUT:
        raise HTTPException(
            status_code=422,
            detail=f"Explaining needs at least {MIN_ROWS_FOR_HOLDOUT} labeled rows.",
        )
    if len(req.feature_columns) > 15:
        raise HTTPException(
            status_code=422,
            detail="Explaining is limited to 15 input columns (one model run per column).",
        )

    stratify = y if task == "classification" and y.value_counts().min() >= 2 else None
    X_train, X_val, y_train, y_val = train_test_split(
        X, y, test_size=HOLDOUT_FRACTION, random_state=42, stratify=stratify
    )

    def score(X_v) -> float:
        result = backend.fit_predict(X_train, y_train, X_v, task, ensemble=req.ensemble)
        if task == "classification":
            return float(accuracy_score(y_val, result.predictions))
        return float(r2_score(y_val, result.predictions))

    try:
        baseline = score(X_val)
        rng = np.random.default_rng(42)
        importances = []
        for feature in req.feature_columns:
            X_perm = X_val.copy()
            X_perm[feature] = rng.permutation(X_perm[feature].values)
            importances.append(
                FeatureImportance(
                    feature=feature, importance=round(baseline - score(X_perm), 4)
                )
            )
    except WeightsUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc

    importances.sort(key=lambda fi: fi.importance, reverse=True)
    return ExplainResponse(
        task=task,
        model_name=getattr(backend, "name", "model"),
        metric_name="accuracy" if task == "classification" else "R²",
        baseline_score=round(baseline, 4),
        n_holdout=len(y_val),
        importances=importances,
    )


# --------------------------------------------------------------------------
# Explain a single row ("why this prediction?")
# --------------------------------------------------------------------------


def _fmt_pred(v) -> str:
    return f"{v:g}" if isinstance(v, (float, np.floating)) else str(v)


@app.post("/api/explain-row", response_model=ExplainRowResponse)
def explain_row(req: ExplainRowRequest) -> ExplainRowResponse:
    predict_req = PredictRequest(
        dataset_id=req.dataset_id,
        target_column=req.target_column,
        feature_columns=req.feature_columns,
        model=req.model,
        task=req.task,
        ensemble=req.ensemble,
    )
    _, _, _, X, y, X_predict, task, backend = _prepare(predict_req)
    if req.row_index not in X_predict.index:
        raise HTTPException(
            status_code=422, detail="That row is not one of the rows being predicted."
        )
    if len(req.feature_columns) > 15:
        raise HTTPException(
            status_code=422,
            detail="Explaining is limited to 15 input columns (one ablation per column).",
        )

    # One batch: the row as-is, then one variant per feature with that value
    # swapped for a "typical" one (median for numbers, most common otherwise).
    row = X_predict.loc[[req.row_index]]
    typicals: dict[str, object] = {}
    variants = [row]
    for feature in req.feature_columns:
        col = X[feature].dropna()
        if pd.api.types.is_numeric_dtype(col):
            typical = col.median()
        else:
            typical = col.mode().iloc[0] if len(col) else None
        typicals[feature] = typical
        variant = row.copy()
        variant[feature] = typical
        variants.append(variant)

    try:
        result = backend.fit_predict(
            X, y, pd.concat(variants, ignore_index=True), task, ensemble=req.ensemble
        )
    except WeightsUnavailable as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc

    base = result.predictions[0]
    contributions = []
    for i, feature in enumerate(req.feature_columns):
        swapped = result.predictions[i + 1]
        if task == "regression":
            impact = abs(float(swapped) - float(base))
        else:
            impact = 0.0 if str(swapped) == str(base) else 1.0
        actual = row[feature].iloc[0]
        contributions.append(
            RowContribution(
                feature=feature,
                value=None if pd.isna(actual) else str(actual),
                typical=_fmt_pred(typicals[feature]) if typicals[feature] is not None else "—",
                prediction=_fmt_pred(swapped),
                impact=round(impact, 4),
            )
        )
    contributions.sort(key=lambda c: c.impact, reverse=True)
    return ExplainRowResponse(
        row_index=req.row_index,
        task=task,
        model_name=getattr(backend, "name", "model"),
        prediction=_fmt_pred(base),
        contributions=contributions,
    )


# --------------------------------------------------------------------------
# Result downloads
# --------------------------------------------------------------------------


@app.get("/api/results/{prediction_id}/csv")
def download_result_csv(prediction_id: str) -> StreamingResponse:
    stored = store.load_result(prediction_id)
    if stored is None:
        raise HTTPException(status_code=404, detail="Result not found.")
    return StreamingResponse(
        io.BytesIO(stored["csv"]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=predictions.csv"},
    )


@app.get("/api/results/{prediction_id}/xlsx")
def download_result_xlsx(prediction_id: str) -> Response:
    """The original workbook with predicted cells filled in and highlighted."""
    stored = store.load_result(prediction_id)
    if stored is None:
        raise HTTPException(status_code=404, detail="Result not found.")
    try:
        ds = datasets.get_dataset(stored["project_id"])
    except KeyError:
        raise HTTPException(status_code=404, detail="The source file is gone; use the CSV download.")

    sheet_name = stored["sheet_name"]
    sheet = ds.sheets.get(sheet_name)
    if sheet is None:
        raise HTTPException(status_code=404, detail="The source sheet is gone; use the CSV download.")

    target_col_idx = list(sheet.df.columns).index(stored["target_column"])
    predicted = {
        row["row_index"]: (row["prediction"], row.get("confidence"))
        for row in stored["response"]["predictions"]
    }

    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill

    fill = PatternFill(start_color="D6E7FA", end_color="D6E7FA", fill_type="solid")
    font = Font(color="1C5CAB", bold=True)

    wb = Workbook()
    wb.remove(wb.active)
    for name, sh in ds.sheets.items():
        ws = wb.create_sheet(title=name[:31])
        for r, row in enumerate(sh.raw.values.tolist()):
            for c, val in enumerate(row):
                ws.cell(row=r + 1, column=c + 1, value=_excel_value(val))
        if name == sheet_name:
            for raw_row, (value, confidence) in predicted.items():
                cell = ws.cell(row=raw_row + 1, column=target_col_idx + 1)
                cell.value = _excel_value(value)
                cell.fill = fill
                cell.font = font
                note = "Predicted by " + stored["response"]["model_name"]
                if confidence is not None:
                    note += f" ({confidence * 100:.0f}% confidence)"
                cell.comment = Comment(note, "TabFM Studio")

    buf = io.BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=predictions.xlsx"},
    )


def _excel_value(val: str):
    if val is None or val == "":
        return None
    try:
        num = float(val)
        return int(num) if num.is_integer() else num
    except (TypeError, ValueError):
        return val


# --------------------------------------------------------------------------
# Metrics helpers
# --------------------------------------------------------------------------


def _label_sort_key(label: str) -> tuple:
    """Numeric class labels order by value ("2" before "10"), text ones alphabetically."""
    try:
        return (0, float(label), label)
    except ValueError:
        return (1, 0.0, label)


def _holdout_metrics(
    backend, X, y, task, warnings: list[str], ensemble: bool = False
) -> Metrics | None:
    """Evaluate on a held-out slice of the labeled rows so users can judge quality."""
    if len(X) < MIN_ROWS_FOR_HOLDOUT:
        warnings.append(
            f"Fewer than {MIN_ROWS_FOR_HOLDOUT} labeled rows; skipping the accuracy check."
        )
        return None
    stratify = y if task == "classification" and y.value_counts().min() >= 2 else None
    X_train, X_val, y_train, y_val = train_test_split(
        X, y, test_size=HOLDOUT_FRACTION, random_state=42, stratify=stratify
    )
    if len(y_val) < MIN_RELIABLE_HOLDOUT:
        warnings.append(
            f"Only {len(y_val)} rows could be held out for the accuracy check, "
            "so its numbers are a rough signal, not a precise score."
        )
    try:
        result = backend.fit_predict(X_train, y_train, X_val, task, ensemble=ensemble)
    except WeightsUnavailable:
        raise
    except Exception:
        logger.exception("Holdout evaluation failed")
        warnings.append("The accuracy check failed; predictions are still produced.")
        return None
    if task == "classification":
        actual = y_val.astype(str)
        predicted = pd.Series(result.predictions).astype(str)
        labels = sorted(set(actual) | set(predicted), key=_label_sort_key)
        matrix = confusion_matrix(actual, predicted, labels=labels)
        return Metrics(
            task=task,
            n_holdout=len(y_val),
            accuracy=round(float(accuracy_score(y_val, result.predictions)), 4),
            f1_macro=round(float(f1_score(y_val, result.predictions, average="macro")), 4),
            confusion=ConfusionMatrix(labels=labels, matrix=matrix.tolist()),
        )
    actual = pd.to_numeric(y_val, errors="coerce")
    predicted = pd.to_numeric(pd.Series(result.predictions), errors="coerce")
    samples = [
        HoldoutSample(actual=round(float(a), 6), predicted=round(float(p), 6))
        for a, p in list(zip(actual, predicted))[:MAX_HOLDOUT_SAMPLES]
        if pd.notna(a) and pd.notna(p)
    ]
    return Metrics(
        task=task,
        n_holdout=len(y_val),
        r2=round(float(r2_score(y_val, result.predictions)), 4),
        mae=round(float(mean_absolute_error(y_val, result.predictions)), 4),
        holdout_samples=samples,
    )


def _prediction_rows(df, labeled_mask, result) -> list[PredictionRow]:
    predict_indices = df.index[~labeled_mask]
    rows = []
    for i, idx in enumerate(predict_indices):
        row = df.loc[idx]
        pred = result.predictions[i]
        rows.append(
            PredictionRow(
                row_index=int(idx),
                values={c: (None if pd.isna(v) else str(v)) for c, v in row.items()},
                prediction=f"{pred:g}" if isinstance(pred, (float, np.floating)) else str(pred),
                confidence=(
                    round(float(result.confidences[i]), 3)
                    if result.confidences is not None
                    else None
                ),
            )
        )
    return rows


def _distribution(predictions: np.ndarray, task) -> list[DistributionBin]:
    if task == "classification":
        counts = pd.Series(predictions).value_counts()
        return [DistributionBin(label=str(k), count=int(v)) for k, v in counts.items()]
    values = pd.to_numeric(pd.Series(predictions), errors="coerce").dropna()
    if values.empty:
        return []
    bins = min(10, max(1, values.nunique()))
    binned = pd.cut(values, bins=bins)

    def fmt(x: float) -> str:
        return f"{x:,.0f}" if abs(x) >= 1000 else f"{x:.4g}"

    return [
        DistributionBin(label=f"{fmt(iv.left)} – {fmt(iv.right)}", count=int(n))
        for iv, n in binned.value_counts().sort_index().items()
    ]


# --------------------------------------------------------------------------
# Static frontend (Docker / single-container deployments)
# --------------------------------------------------------------------------

_frontend_dist = os.environ.get("FRONTEND_DIST")
if _frontend_dist and os.path.isdir(_frontend_dist):
    from fastapi.staticfiles import StaticFiles

    # Mounted last so all /api routes above keep precedence.
    app.mount("/", StaticFiles(directory=_frontend_dist, html=True), name="frontend")
